from datetime import datetime
from django.utils.html import format_html
from django.contrib import admin, messages
from django.http import JsonResponse, HttpResponse
from simpleui.admin import AjaxAdmin
from .models import BranchStore, CashRedemption, CashExchange, SalesInviteCode
from redeem.models import Redeem
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

@admin.register(BranchStore)
class BranchStoreAdmin(admin.ModelAdmin):
    list_display = ('employee', 'consumer', 'name', 'address', 'status')
    list_filter = ['status']
    search_fields = ('consumer__mobile', 'employee')
    actions = ['set_staff', 'reject_staff']
    autocomplete_fields = ('consumer', )
    list_per_page = 20
    # fields = ['status', 'consumer', 'employee', 'name', 'address']
    fieldsets = (
        ('审核信息', {
            'fields': ('status',)
        }),
        ('员工信息', {
            'fields': ('consumer', 'employee')
        }),
        ('门店信息', {
            'fields': ('name', 'address')
        }),
    )

    def get_list_per_page(self, request):
        try:
            return int(request.GET.get('per_page', self.list_per_page))
        except ValueError:
            return self.list_per_page

    def changelist_view(self, request, extra_context=None):
        per_page = self.get_list_per_page(request)
        extra_context = extra_context or {}
        extra_context.update({
            'per_page_options': [10, 20, 50, 100],
            'per_page': per_page,
        })
        self.list_per_page = per_page  # 动态设置
        return super().changelist_view(request, extra_context=extra_context)

    def set_staff(self, request, queryset):
        queryset.update(status=True)
        self.message_user(request, "操作成功！", level=messages.SUCCESS)

    set_staff.short_description = ' 通过审核'
    set_staff.type = 'warning'
    set_staff.icon = 'fa-solid fa-face-grin'
    set_staff.enable = True

    def reject_staff(self, request, queryset):
        queryset.filter(status=False).delete()
        self.message_user(request, "操作成功！", level=messages.SUCCESS)

    reject_staff.short_description = ' 驳回资料'
    reject_staff.type = 'danger'
    reject_staff.icon = 'fa-solid fa-face-frown-open'
    reject_staff.enable = True

    # autocomplete_fields = ('consumer',)


@admin.register(CashRedemption)
class CashRedemptionAdmin(AjaxAdmin):
    list_display = ('number', 'redeem', 'cash', 'status', 'updated_at')
    search_fields = ('redeem__number', 'number')
    list_per_page = 20
    actions = ['bulk_generate', 'export_data']
    fieldsets = (
        ('抽奖码信息', {
            'fields': ('redeem',)
        }),
        ('基础信息', {
            'fields': ('status', 'number', 'cash')
        }),
    )
    autocomplete_fields = ('redeem',)

    def get_list_per_page(self, request):
        try:
            return int(request.GET.get('per_page', self.list_per_page))
        except ValueError:
            return self.list_per_page

    def changelist_view(self, request, extra_context=None):
        per_page = self.get_list_per_page(request)
        extra_context = extra_context or {}
        extra_context.update({
            'per_page_options': [10, 20, 50, 100],
            'per_page': per_page,
        })
        self.list_per_page = per_page  # 动态设置
        return super().changelist_view(request, extra_context=extra_context)

    def bulk_generate(self, request, queryset):
        total = request.POST.get('count', 0)
        redeem_list = []
        cash = request.POST.get('cash', 1.0)
        prefix = request.POST.get('prefix', None)
        if prefix:
            if len(prefix) < 4:
                return JsonResponse(data={
                    'status': 'error',
                    'msg': '前缀系列CODE长度不足4个字符'
                })
            if len(prefix) > 15:
                return JsonResponse(data={
                    'status': 'error',
                    'msg': '前缀系列ODE长度最大不超过15个字符'
                })
            for i in range(0, int(total)):
                redeem_list.append(CashRedemption(number=CashRedemption.generate_prefix(prefix), cash=cash))
        else:
            for i in range(0, int(total)):
                redeem_list.append(CashRedemption(number=CashRedemption.generate_number(), cash=cash))
        CashRedemption.objects.bulk_create(redeem_list)
        return JsonResponse(data={
            'status': 'success',
            'msg': f'批量生成 {len(redeem_list)} 现金兑换码',
        })

    bulk_generate.short_description = '  批量生成'
    bulk_generate.type = 'success'
    bulk_generate.icon = 'fa-solid fa-bowl-rice'

    bulk_generate.layer = {
        # 弹出层中的输入框配置

        # 这里指定对话框的标题
        'title': '批量生成兑换码',
        # 提示信息
        'tips': '现金兑换码格式为 XUK9-CRPM-VX4Y-F1AH ，由4部分组成，每部分为4个字符，'
                '可以指定第一部分字符串为统一内容前缀。如不指定前缀，请留空，则随机生成。',
        # 确认按钮显示文本
        'confirm_button': '提交',
        # 取消按钮显示文本
        'cancel_button': '取消',

        # 弹出层对话框的宽度，默认50%
        'width': '25%',

        # 表单中 label的宽度，对应element-ui的 label-width，默认80px
        'labelWidth': "80px",
        'params': [{
            'type': 'input',
            'key': 'prefix',
            'size': 'small',
            'width': '200px',
            'label': '前缀'
        }, {
            'type': 'input',
            'key': 'count',
            'size': 'small',
            'width': '200px',
            'label': '数量',
            'require': True
        }, {
            'type': 'input',
            'key': 'cash',
            'size': 'small',
            'width': '200px',
            'label': '现金',
            'require': True
        }]
    }

    def export_data(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = '现金兑换码列表'

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        headers = ['现金兑换码', '关联抽奖码', '兑换金额', '状态', '创建时间']
        for i, header in enumerate(headers):
            ws.cell(row=1, column=i + 1, value=header)
            ws.cell(row=1, column=i + 1).font = header_font
            ws.cell(row=1, column=i + 1).fill = header_fill
            ws.cell(row=1, column=i + 1).alignment = Alignment(horizontal='center')

        for obj in queryset:
            row = [obj.number, obj.redeem_number(), obj.cash, obj.status, obj.created_at.strftime('%Y-%m-%d %H:%M:%S')]
            ws.append(row)

        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 10

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        filename = f"现金兑换码_{timestamp}_数据.xlsx"
        response["Content-Disposition"] = f'attachment; filename*=UTF-8 {filename}'
        wb.save(response)
        return response

    export_data.short_description = ' 导出'
    export_data.type = 'warning'
    export_data.icon = 'fa fa-solid fa-file-export'


@admin.register(CashExchange)
class CashExchangeAdmin(admin.ModelAdmin):
    list_filter = ('status',)
    search_fields = ['user__mobile', 'redemption__number']
    list_display = ['staff_info', 'number', 'cash', 'status', 'withdrawal_at']
    autocomplete_fields = ('user', 'redemption')
    list_per_page = 20
    fieldsets = (
        ('用户信息', {
            'fields': ('user', )
        }),
        ('兑换码信息', {
            'fields': ('number', 'redemption', 'cash', 'status')
        }),
    )

    def get_list_per_page(self, request):
        try:
            return int(request.GET.get('per_page', self.list_per_page))
        except ValueError:
            return self.list_per_page

    def changelist_view(self, request, extra_context=None):
        per_page = self.get_list_per_page(request)
        extra_context = extra_context or {}
        extra_context.update({
            'per_page_options': [10, 20, 50, 100],
            'per_page': per_page,
        })
        self.list_per_page = per_page  # 动态设置
        return super().changelist_view(request, extra_context=extra_context)

    def staff_info(self, obj):
        if obj.user.is_sales():
            try:
                staff_store = obj.user.branch_store
                return f"{staff_store.name} - {staff_store.employee} - {obj.user.mobile}"
            except Exception as e:
                return obj.user
        else:
            return obj.user.mobile

    staff_info.short_description = '员工信息'


@admin.register(SalesInviteCode)
class SalesInviteCodeAdmin(AjaxAdmin):
    list_display = [
        "code",
        "is_used",
        "created_at",
        "copy_code",
    ]
    list_filter = ["is_used"]
    search_fields = ["code"]
    readonly_fields = ["code", "created_at"]
    actions = ["bulk_generate"]
    list_per_page = 20
    fieldsets = (
        ('邀请码', {
            'fields': ('code', "is_used")
        }),
    )

    def get_list_per_page(self, request):
        try:
            return int(request.GET.get('per_page', self.list_per_page))
        except ValueError:
            return self.list_per_page

    def changelist_view(self, request, extra_context=None):
        per_page = self.get_list_per_page(request)
        extra_context = extra_context or {}
        extra_context.update({
            'per_page_options': [10, 20, 50, 100],
            'per_page': per_page,
        })
        self.list_per_page = per_page  # 动态设置
        return super().changelist_view(request, extra_context=extra_context)

    def copy_code(self, obj):
        invite_url = f"https://fuxion.fun/employee/invite/{obj.code}"
        if not obj.is_used:
            return format_html(
                """
                <button class="el-button el-button--primary el-button--small" onclick="
                    event.preventDefault();
                    var text = '{invite_url}';
                    if (navigator.clipboard) {{
                        navigator.clipboard.writeText(text).then(() => alert('已复制邀请链接！\\n' + text));
                    }}
                ">复制邀请链接</button>
                """,
                invite_url=invite_url
            )
        else:
            return '已使用'

    copy_code.short_description = " 操作"

    # 隐藏默认的“增加”按钮，只允许列表页生成
    def has_add_permission(self, request):
        return False

    def bulk_generate(self, request, queryset):
        total = request.POST.get('count', 0)
        code_list = []
        for i in range(0, int(total)):
            code_list.append(SalesInviteCode(code=SalesInviteCode.generate_code()))
        SalesInviteCode.objects.bulk_create(code_list)
        return JsonResponse(data={
            'status': 'success',
            'msg': f'批量生成 {len(code_list)} 邀请码',
        })

    bulk_generate.short_description = '  批量生成'
    bulk_generate.type = 'success'
    bulk_generate.icon = 'fa-solid fa-fill-drip'

    bulk_generate.layer = {
        # 弹出层中的输入框配置

        # 这里指定对话框的标题
        'title': '批量生成邀请码',
        # 提示信息
        'tips': '邀请用户成为销售角色，用以参与商品销售现金奖励，增加员工被动收入，调动主观积极性'
                '更好服务于消费者，促进商品成交率',
        # 确认按钮显示文本
        'confirm_button': '提交',
        # 取消按钮显示文本
        'cancel_button': '取消',

        # 弹出层对话框的宽度，默认50%
        'width': '25%',

        # 表单中 label的宽度，对应element-ui的 label-width，默认80px
        'labelWidth': "80px",
        'params': [{
            'type': 'input',
            'key': 'count',
            'size': 'small',
            'width': '200px',
            'label': '数量',
            'require': True
        }]
    }

