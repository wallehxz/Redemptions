import os
from django.contrib import admin
from django.http import JsonResponse, HttpResponse
from openpyxl import load_workbook
from simpleui.admin import AjaxAdmin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from datetime import datetime

from series.models import Series, Prize
from .models import Redeem, Redemption
from django.core.files.storage import FileSystemStorage


@admin.register(Redeem)
class RedeemAdmin(AjaxAdmin):
    list_display = ('number', 'prize', 'series_name', 'status', 'created_at')
    fields = ('series', 'prize', 'status', 'number')
    list_filter = ['prize', 'status', 'series']
    search_fields = ['number']
    list_per_page = 20
    # actions = ['bulk_generate', 'export_data']
    actions = ['export_data', 'bulk_generate', 'import_codes']

    def series_name(self, obj):
        if obj.prize and obj.prize.series:
            return obj.prize.series.name
        if obj.series:
            return obj.series.name
        else:
            return '--'

    series_name.short_description = '所属系列'

    def export_data(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = '兑换码列表'

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        headers = ['兑换码', '奖品名称', '所属系列', '状态', '创建时间']
        for i, header in enumerate(headers):
            ws.cell(row=1, column=i + 1, value=header)
            ws.cell(row=1, column=i + 1).font = header_font
            ws.cell(row=1, column=i + 1).fill = header_fill
            ws.cell(row=1, column=i + 1).alignment = Alignment(horizontal='center')

        for obj in queryset:
            row = [obj.number, obj.prize.name, obj.prize.series.name, obj.get_status_display(),
                   obj.created_at.strftime('%Y-%m-%d %H:%M:%S')]
            ws.append(row)

        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 10

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        filename = f"兑换码_{timestamp}_数据.xlsx"
        response["Content-Disposition"] = f'attachment; filename*=UTF-8 ''{filename}'
        wb.save(response)
        return response

    export_data.short_description = ' 导出'
    export_data.type = 'warning'
    export_data.icon = 'fa fa-solid fa-file-export'

    def bulk_generate(self, request, queryset):
        total = request.POST.get('count', 0)
        redeem_list = []
        prefix = request.POST.get('prefix', None)
        if prefix:
            if len(prefix) < 4:
                return JsonResponse(data={
                    'status': 'error',
                    'msg': '前缀系列CODE长度不足4个字符'
                })
            if len(prefix) > 15:
                return JsonResponse(data={
                    'status': 'error',
                    'msg': '前缀系列ODE长度最大不超过15个字符'
                })
            for i in range(0, int(total)):
                # redeem_list.append(Redeem(number=Redeem.generate_prefix(prefix), prize_id=request.POST.get('prize')))
                redeem_list.append(Redeem(number=Redeem.generate_prefix(prefix)))
        else:
            for i in range(0, int(total)):
                redeem_list.append(Redeem(number=Redeem.generate_number(), prize_id=request.POST.get('prize')))
        Redeem.objects.bulk_create(redeem_list)
        return JsonResponse(data={
            'status': 'success',
            'msg': f'批量生成 {len(redeem_list)} 抽奖码',
        })

    bulk_generate.short_description = ' 批量'
    bulk_generate.type = 'success'
    bulk_generate.icon = 'fa fa-solid fa-bowl-rice'

    bulk_generate.layer = {
        # 弹出层中的输入框配置

        # 这里指定对话框的标题
        'title': '批量生成兑换码',
        # 提示信息
        'tips': '兑换码格式为 XUK9-CRPM-VX4Y-F1AH ，由4部分组成，每部分为4个字符，'
                '可以指定第一部分字符串为统一内容前缀。如不指定前缀，请留空，则随机生成。',
        # 确认按钮显示文本
        'confirm_button': '提交',
        # 取消按钮显示文本
        'cancel_button': '取消',

        # 弹出层对话框的宽度，默认50%
        'width': '30%',

        # 表单中 label的宽度，对应element-ui的 label-width，默认80px
        'labelWidth': "80px",
        'params': [{
            'type': 'input',
            'key': 'prefix',
            'size': 'small',
            'width': '300px',
            'label': '前缀'
        }, {
            'type': 'input',
            'key': 'count',
            'size': 'small',
            'width': '300px',
            'label': '数量',
            'require': True
        }]
        #{
        #     'type': 'select',
        #     'key': 'prize',
        #     'label': '奖品',
        #     'width': '300px',
        #     'size': 'small',
        #     'value': '',
        #     'require': True,
        #     'options': Redeem.prize_list(),
        # }]
    }

    def import_codes(self, request, queryset):
        xlsx_file = request.FILES['upload']
        fs = FileSystemStorage()
        filename = fs.save(xlsx_file.name, xlsx_file)
        file_path = fs.path(filename)
        try:
            # 加载Excel文件
            wb = load_workbook(filename=file_path)
            code_total = 0
            for sheet_name in wb.sheetnames:
                code_list = []
                sheet = wb[sheet_name]
                print(f"processing sheet {sheet_name} rows data")
                for row in sheet.iter_rows(values_only=True):
                    if row[2] == '未被扫描':
                        existing_codes = set(Redeem.objects.values_list('number', flat=True))
                        if row[0] in existing_codes:
                            print(f"{row[0]} existing，skip")
                            continue
                        series = None
                        prize = None
                        if row[1]:
                            series, _ = Series.objects.get_or_create(name=row[1])
                        if row[3]:
                            prize, _ = Prize.objects.get_or_create(name=row[3], series=series)
                        code = Redeem(number=row[0], series=series, prize=prize)
                        code_list.append(code)
                code_total += len(code_list)
                Redeem.objects.bulk_create(code_list)
            # 处理完成后删除临时文件
            os.remove(file_path)
            return JsonResponse(data={
                'status': 'success',
                'msg': f'导入 {len(wb.sheetnames)} sheets {code_total} 抽奖码',
            })
        except Exception as e:
            return JsonResponse(data={
                'status': 'error',
                'msg': f"{e}"
            })

    import_codes.short_description = '导入'
    import_codes.type = 'danger'
    import_codes.icon = 'el-icon-upload'
    import_codes.enable = True

    import_codes.layer = {
        'title': '抽奖码导入',
        'tips': '请勿重复导入抽奖码',
        'width': '500px',
        'params': [{
            'type': 'file',
            'key': 'upload',
            'label': '文件'
        }]
    }


@admin.register(Redemption)
class RedemptionAdmin(admin.ModelAdmin):
    list_display = ('prize', 'redeem', 'consumer', 'shipping', 'status', 'express_info', 'created_at')
    fields = ['status', 'express_name', 'express_order']
    # raw_id_fields = ['redeem']
    # autocomplete_fields = ['redeem']
    search_fields = ['redeem__number']
    list_filter = ['status']

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.filter(prize__isnull=False)

    def express_info(self, obj):
        return f'{obj.express_name}-{obj.express_order}'

    express_info.short_description = '快递信息'


# Register your models here.
