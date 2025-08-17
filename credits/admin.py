from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from django.contrib import admin
from simpleui.admin import AjaxAdmin
from django.http import JsonResponse, HttpResponse
from django.utils.html import format_html
import os
from .models import Product, ProductSpec, PointsTransaction, RedemptionCode, ExchangeOrder
from django.core.files.storage import FileSystemStorage
from openpyxl import load_workbook


class ProductSpecInline(admin.TabularInline):
    model = ProductSpec
    extra = 0
    min_num = 1
    fields = ['name', 'points_required', 'stock', 'is_default']
    
    # 自定义列标题
    verbose_name = "商品规格"
    verbose_name_plural = "商品规格列表"
    
    # 设置列宽度和对齐
    classes = ['tabular']

    class Media:
        css = {
            'all': ('stylesheets/product_spec_fix.css',)
        }


@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    list_display = ('name', 'main_img', 'specs_list', 'created_at')
    search_fields = ['name']
    list_filter = ['specs__points_required']
    fieldsets = (
        ('基本信息', {
            'fields': ('name', 'points_required', 'main_image')
        }),
        ('商品详情', {
            'fields': ('description',),
            'description': '如有多张图推荐使用 <a href="https://fulicat.com/lab/pintu/" target="_blank">在线工具</a>  进行拼接<br>'
        }),
    )
    inlines = [ProductSpecInline]

    list_per_page = 10

    def get_list_per_page(self, request):
        try:
            return int(request.GET.get('per_page', self.list_per_page))
        except ValueError:
            return self.list_per_page

    def changelist_view(self, request, extra_context=None):
        per_page = self.get_list_per_page(request)
        extra_context = extra_context or {}
        extra_context.update({
            'per_page_options': [10, 20, 50, 100],
            'per_page': per_page,
        })
        self.list_per_page = per_page  # 动态设置
        return super().changelist_view(request, extra_context=extra_context)

    def specs_list(self, obj):
        specs_str = ""
        for spec in obj.specs.all():
            specs_str += f'规格：{spec.name} 积分：{spec.points_required} 库存：{spec.stock}<br>'
        return format_html(specs_str)

    specs_list.short_description = '规格列表'

    def main_img(self, obj):
        if obj.main_image:
            return format_html('<img src="{}" width="75" height="75" />', obj.main_image.url)

    main_img.short_description = '封面'


@admin.register(RedemptionCode)
class RedemptionCodeAdmin(AjaxAdmin):
    list_display = ('code', 'is_used', 'points_value', 'used_by', 'used_at')
    fields = ['code', 'points_value', 'is_used', 'used_by']
    search_fields = ['code']
    list_filter = ['used_by', 'is_used']
    list_per_page = 20
    actions = ['bulk_generate', 'import_codes']

    def get_list_per_page(self, request):
        try:
            return int(request.GET.get('per_page', self.list_per_page))
        except ValueError:
            return self.list_per_page

    def changelist_view(self, request, extra_context=None):
        per_page = self.get_list_per_page(request)
        extra_context = extra_context or {}
        extra_context.update({
            'per_page_options': [10, 20, 50, 100],
            'per_page': per_page,
        })
        self.list_per_page = per_page  # 动态设置
        return super().changelist_view(request, extra_context=extra_context)

    def bulk_generate(self, request, queryset):
        total = request.POST.get('total', 0)
        points_value = request.POST.get('points_value', 1)
        redeem_list = []
        prefix = request.POST.get('prefix', None)
        if prefix and len(prefix) < 4:
            return JsonResponse(data={
                'status': 'error',
                'msg': '自定义前缀长度不足4个字符'
            })
        for i in range(0, int(total)):
            if prefix:
                redeem_list.append(RedemptionCode(code=RedemptionCode.generate_prefix(prefix), points_value=points_value))
            else:
                redeem_list.append(RedemptionCode(code=RedemptionCode.generate_code(), points_value=points_value))
        RedemptionCode.objects.bulk_create(redeem_list)
        tip_msg = f'批量生产兑换码 {total} 条'
        return JsonResponse(data={
            'status': 'success',
            'msg': tip_msg
        })

    bulk_generate.short_description = ' 批量'
    bulk_generate.type = 'success'
    bulk_generate.icon = 'fa fa-solid fa-hands-bubbles'

    bulk_generate.layer = {
        # 弹出层中的输入框配置

        # 这里指定对话框的标题
        'title': '批量生成兑换码',
        # 提示信息
        'tips': '兑换码用于兑换商品，默认兑换的福力为 1 ;'
                '兑换码的格式AAAA-BBBB-CCCC-DDDD ;'
                '可以自定义开头前缀部分，不填则随机生成',
        # 确认按钮显示文本
        'confirm_button': '提交',
        # 取消按钮显示文本
        'cancel_button': '取消',

        # 弹出层对话框的宽度，默认50%
        'width': '30%',

        # 表单中 label的宽度，对应element-ui的 label-width，默认80px
        'labelWidth': "80px",
        'params': [{
            'type': 'input',
            'key': 'total',
            'size': 'small',
            'width': '300px',
            'label': '数量',
            'require': True
        },{
            'type': 'input',
            'key': 'points_value',
            'size': 'small',
            'width': '300px',
            'label': '积分值',
            'value': '1',
            'require': True
        },{
            'type': 'input',
            'key': 'prefix',
            'size': 'small',
            'width': '300px',
            'label': '自定前缀',
            'value': '',
            'require': False
        }]
        #{
        #     'type': 'select',
        #     'key': 'prize',
        #     'label': '奖品',
        #     'width': '300px',
        #     'size': 'small',
        #     'value': '',
        #     'require': True,
        #     'options': Redeem.prize_list(),
        # }]

    }

    def import_codes(self, request, queryset):
        xlsx_file = request.FILES['upload']
        fs = FileSystemStorage()
        filename = fs.save(xlsx_file.name, xlsx_file)
        file_path = fs.path(filename)
        try:
            # 加载Excel文件
            wb = load_workbook(filename=file_path)
            code_total = 0
            for sheet_name in wb.sheetnames:
                code_list = []
                sheet = wb[sheet_name]
                print(f"processing sheet {sheet_name} rows data")
                for row in sheet.iter_rows(values_only=True):
                    if row[2] == '未被扫描':
                        existing_codes = set(RedemptionCode.objects.values_list('code', flat=True))
                        if row[0] in existing_codes:
                            print(f"{row[0]} existing，skip")
                            continue
                        points_value = 1
                        if row[3] and int(row[3]) > 0:
                            points_value = int(row[3])
                        r_code = RedemptionCode(code=row[0], points_value=points_value)
                        code_list.append(r_code)
                code_total += len(code_list)
                RedemptionCode.objects.bulk_create(code_list)
            # 处理完成后删除临时文件
            os.remove(file_path)
            return JsonResponse(data={
                'status': 'success',
                'msg': f'导入 {len(wb.sheetnames)} sheets {code_total} 兑换码',
            })
        except Exception as e:
            return JsonResponse(data={
                'status': 'error',
                'msg': f"{e}"
            })

    import_codes.short_description = '导入'
    import_codes.type = 'danger'
    import_codes.icon = 'el-icon-upload'
    import_codes.enable = True

    import_codes.layer = {
        'title': '兑换码导入',
        'tips': '请勿重复导入兑换码',
        'width': '500px',
        'params': [{
            'type': 'file',
            'key': 'upload',
            'label': '文件'
        }]
    }


@admin.register(PointsTransaction)
class PointsTransactionAdmin(AjaxAdmin):
    list_filter = ('user__mobile',)
    list_display = ('user', 'amount', 'transaction_type', 'description', 'created_at')
    list_per_page = 20

    def get_list_per_page(self, request):
        try:
            return int(request.GET.get('per_page', self.list_per_page))
        except ValueError:
            return self.list_per_page

    def changelist_view(self, request, extra_context=None):
        per_page = self.get_list_per_page(request)
        extra_context = extra_context or {}
        extra_context.update({
            'per_page_options': [10, 20, 50, 100],
            'per_page': per_page,
        })
        self.list_per_page = per_page  # 动态设置
        return super().changelist_view(request, extra_context=extra_context)


@admin.register(ExchangeOrder)
class ExchangeOrderAdmin(AjaxAdmin):
    list_filter = ('status',)
    readonly_fields = ('order_number',)
    search_fields = ['user__mobile', 'tracking_number']
    fields = ('order_number', 'express_name', 'tracking_number', 'note')
    list_display = ('user', 'express', 'status', 'note', 'harvest_info','created_at')
    list_per_page = 20
    actions = ['export_data']

    def get_list_per_page(self, request):
        try:
            return int(request.GET.get('per_page', self.list_per_page))
        except ValueError:
            return self.list_per_page

    def changelist_view(self, request, extra_context=None):
        per_page = self.get_list_per_page(request)
        extra_context = extra_context or {}
        extra_context.update({
            'per_page_options': [10, 20, 50, 100],
            'per_page': per_page,
        })
        self.list_per_page = per_page  # 动态设置
        return super().changelist_view(request, extra_context=extra_context)

    def get_queryset(self, request):
        return super().get_queryset(request).select_related('harvest')

    def express(self, obj):
        if obj.tracking_number:
            return f"{obj.express_name} - {obj.tracking_number}"
        else:
            return "--"
    express.short_description = '快递信息'

    def harvest_info(self, obj):
        if obj.harvest:
            harvest = obj.harvest
            return format_html(
                """
                <span><strong>姓名：</strong> {}<br></span>
                <span><strong>电话：</strong> {}<br></span>
                <span><strong>地址：</strong> {}<br></span>
                """,
                harvest.nick_name,
                harvest.mobile,
                harvest.full_address(),
            )
        else:
            return "--"

    harvest_info.short_description = '收货信息'

    def export_data(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = '兑换码列表'

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        headers = ['收件人', '联系电话', '收获地址', '商品信息', '快递单号', '订单状态', '创建时间']
        for i, header in enumerate(headers):
            ws.cell(row=1, column=i + 1, value=header)
            ws.cell(row=1, column=i + 1).font = header_font
            ws.cell(row=1, column=i + 1).fill = header_fill
            ws.cell(row=1, column=i + 1).alignment = Alignment(horizontal='center')

        for obj in queryset:
            if obj.harvest:
                row = [obj.harvest.nick_name, obj.harvest.mobile, obj.harvest.full_address(), obj.note, obj.tracking_number, obj.get_status_display(),
                       obj.created_at.strftime('%Y-%m-%d %H:%M:%S')]
                ws.append(row)

        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 10

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        filename = f"商品兑换订单_{timestamp}_数据.xlsx"
        response["Content-Disposition"] = f'attachment; filename*=UTF-8 ''{filename}'
        wb.save(response)
        return response

    export_data.short_description = ' 导出'
    export_data.type = 'success'
    export_data.icon = 'fa fa-solid fa-file-export'

