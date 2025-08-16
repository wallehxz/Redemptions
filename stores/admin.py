import os
from time import sleep
from django.contrib import admin, messages
from django.forms import ModelForm, Media
from django.http import JsonResponse
from django.utils.html import format_html
from django.utils.safestring import mark_safe
from django.conf import settings
from simpleui.admin import AjaxAdmin
from django.core.files.storage import FileSystemStorage
from openpyxl import load_workbook
from .models import Store, Plush


class StoreAdminForm(ModelForm):
    """自定义店铺表单，集成高德地图"""

    class Meta:
        model = Store
        fields = '__all__'

    class Media:
        css = {
            'all': ('admin/css/store_admin.css',)
        }
        js = (
            f'https://webapi.amap.com/maps?v=2.0&key={getattr(settings, "AMAP_API_KEY", "")}',
            'admin/js/amap_widget.js',
        )


@admin.register(Store)
class StoreAdmin(AjaxAdmin):
    form = StoreAdminForm
    list_display = ['name', 'address_info', 'geo_info', 'is_active', 'created_at']
    list_filter = ['is_active', 'created_at']
    search_fields = ['name', 'address', 'phone']
    actions = ['import_stores', 'update_navigation']

    fieldsets = (
        ('基本信息', {
            'fields': ('name', 'address', 'navigation')
        }),
        ('位置信息', {
            'fields': ('latitude', 'longitude', 'map_widget'),
            'description': '点击地图选择店铺位置，或输入地址搜索获取坐标'
        }),
        ('营业信息', {
            'fields': ('business_hours', 'is_active', 'phone', 'description')
        }),
    )

    readonly_fields = ['map_widget']

    def map_widget(self, obj):
        """地图插件"""
        return mark_safe(f'''
                <div style="margin: 10px 0;">
                    <input type="text" id="address-input" placeholder="输入地址名称搜索..." 
                           style="width: 400px; padding: 5px; margin-right: 10px;">
                    <button type="button" id="search-btn" style="padding: 5px 10px;">搜索地址</button>
                </div>
                <div id="amap-container" style="width: 750px; height: 400px; margin: 10px 0;"></div>
                <div id="location-info" style="margin: 10px 0; color: #666;"></div>
                ''')

    map_widget.short_description = '地图选择'

    class Media:
        css = {
            'all': ('admin/css/store_admin.css',)
        }
        js = (
            f'https://webapi.amap.com/maps?v=2.0&key={getattr(settings, "AMAP_API_KEY", "")}',
            'admin/js/amap_widget.js',
        )

    def import_stores(self, request, queryset):
        xlsx_file = request.FILES['upload']
        fs = FileSystemStorage()
        filename = fs.save(xlsx_file.name, xlsx_file)
        file_path = fs.path(filename)
        try:
            # 加载Excel文件
            wb = load_workbook(filename=file_path)
            code_total = 0
            for sheet_name in wb.sheetnames:
                store_list = []
                sheet = wb[sheet_name]
                print(f"processing sheet {sheet_name} rows data")
                for row in sheet.iter_rows(values_only=True):
                    if row[1] != '地址信息':
                        sleep(0.35)
                        l_address = row[1].replace(" ", "")
                        geocode = Store.geocode_address(l_address)
                        print(geocode)
                        store = Store(name=row[0], address=l_address, latitude=geocode['latitude'], longitude=geocode['longitude'], business_hours='09:00-22:00')
                        store_list.append(store)
                code_total += len(store_list)
                Store.objects.bulk_create(store_list)
            # 处理完成后删除临时文件
            os.remove(file_path)
            return JsonResponse(data={
                'status': 'success',
                'msg': f'导入 {len(wb.sheetnames)} sheets {code_total} 店铺信息',
            })
        except Exception as e:
            return JsonResponse(data={
                'status': 'error',
                'msg': f"{e}"
            })

    import_stores.short_description = '导入'
    import_stores.type = 'success'
    import_stores.icon = 'el-icon-upload'
    import_stores.enable = True

    import_stores.layer = {
        'title': '店铺地址导入',
        'tips': '请勿重复导入店铺',
        'width': '500px',
        'params': [{
            'type': 'file',
            'key': 'upload',
            'label': '文件'
        }]
    }

    def geo_info(self, obj):
        if obj.latitude:
            return format_html(
                """
                <span><strong>{},{}</span>
                """,
                obj.latitude,
                obj.longitude,
            )
        else:
            return "--"

    geo_info.short_description = '地理经纬'

    def address_info(self, obj):
        if obj.latitude:
            return format_html(
                """
                <span><strong>{}<br>{}</span>
                """,
                obj.navigation,
                obj.address,
            )
        else:
            return "--"

    address_info.short_description = '地址信息'

    def update_navigation(self, request, queryset):
        try:
            Store.update_all_navigation_with_address()
            self.message_user(request, "操作成功！", level=messages.SUCCESS)
        except Exception as e:
            self.message_user(request, "操作失败！", level=messages.SUCCESS)

    update_navigation.short_description = ' 导航地址'
    update_navigation.type = 'danger'
    update_navigation.icon = 'fa-solid fa-arrows-rotate'


@admin.register(Plush)
class PlushAdmin(admin.ModelAdmin):
    list_display = ['name','main_img', 'is_latest', 'created_at']
    list_filter = ['is_latest']
    search_fields = ['name']

    fieldsets = (
        ('基本信息', {
            'fields': ('is_latest', 'name', 'main_image')
        }),
        ('周边详情', {
            'fields': ('description',),
            'description': '如有多张图推荐使用 <a href="https://fulicat.com/lab/pintu/" target="_blank">在线工具</a>  进行拼接<br>'
        }),
    )

    def main_img(self, obj):
        if obj.main_image:
            return format_html('<img src="{}" width="auto" height="75" />', obj.main_image.url)

    main_img.short_description = '头图'
